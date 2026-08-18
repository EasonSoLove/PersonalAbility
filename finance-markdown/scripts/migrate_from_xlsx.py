from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from common import code6, decimal_text, dump_funds_yaml, iso_date, load_schema, write_csv_atomic, write_text_atomic

FUND_HEADERS = ["基金代码", "基金名称", "主题", "跟踪指数", "关联场内ETF代码", "关联场内ETF名称", "交易所", "行情SecID", "成立日期", "C类销售服务费/年", "赎回费及说明", "天天基金页面", "信息同步日期"]
TX_HEADERS = ["申请日期", "确认日期", "基金代码", "基金名称(自动)", "交易类型", "申请/成交金额", "确认份额", "确认净值", "手续费", "实际到账/分红", "现金流(自动)", "策略标签", "备注"]


def header_map(ws: Any, row: int) -> dict[str, int]:
    return {str(cell.value).strip(): cell.column for cell in ws[row] if cell.value not in (None, "")}


def get_value(ws: Any, row: int, mapping: dict[str, int], name: str) -> Any:
    col = mapping.get(name)
    return ws.cell(row=row, column=col).value if col else None


def ensure_headers(mapping: dict[str, int], expected: list[str], sheet: str) -> None:
    missing = [x for x in expected if x not in mapping]
    if missing:
        raise ValueError(f"{sheet}缺少字段：{missing}")


def make_tx_id(request_date: str, code: str, counters: dict[tuple[str, str], int]) -> str:
    day = request_date.replace("-", "")
    key = (day, code)
    counters[key] += 1
    return f"TX-{day}-{code}-{counters[key]:03d}"


def migrate(workbook: Path, target_root: Path, force: bool = False) -> dict[str, Any]:
    funds_path = target_root / "data" / "funds.yaml"
    tx_path = target_root / "data" / "transactions.csv"
    if not force and (funds_path.exists() or tx_path.exists()):
        raise FileExistsError("目标数据文件已存在；如需重新迁移请显式使用--force")

    wb = load_workbook(workbook, data_only=True, read_only=False)
    if "基金信息" not in wb.sheetnames or "交易流水" not in wb.sheetnames:
        raise ValueError("工作簿必须包含‘基金信息’和‘交易流水’")

    fund_ws = wb["基金信息"]
    fund_hm = header_map(fund_ws, 5)
    ensure_headers(fund_hm, FUND_HEADERS, "基金信息")
    funds: list[dict[str, Any]] = []
    for row in range(6, fund_ws.max_row + 1):
        code = code6(get_value(fund_ws, row, fund_hm, "基金代码"))
        if not code:
            continue
        etf_code = code6(get_value(fund_ws, row, fund_hm, "关联场内ETF代码")) if get_value(fund_ws, row, fund_hm, "关联场内ETF代码") not in (None, "") else ""
        funds.append({
            "code": code,
            "name": str(get_value(fund_ws, row, fund_hm, "基金名称") or "").strip(),
            "theme": str(get_value(fund_ws, row, fund_hm, "主题") or "").strip(),
            "tracked_index": str(get_value(fund_ws, row, fund_hm, "跟踪指数") or "").strip(),
            "linked_etf": {
                "code": etf_code,
                "name": str(get_value(fund_ws, row, fund_hm, "关联场内ETF名称") or "").strip(),
                "exchange": str(get_value(fund_ws, row, fund_hm, "交易所") or "").strip(),
                "secid": str(get_value(fund_ws, row, fund_hm, "行情SecID") or "").strip(),
            },
            "inception_date": iso_date(get_value(fund_ws, row, fund_hm, "成立日期")),
            "sales_service_fee": float(get_value(fund_ws, row, fund_hm, "C类销售服务费/年")) if get_value(fund_ws, row, fund_hm, "C类销售服务费/年") not in (None, "") else None,
            "redemption_note": str(get_value(fund_ws, row, fund_hm, "赎回费及说明") or "").strip(),
            "source_url": str(get_value(fund_ws, row, fund_hm, "天天基金页面") or "").strip(),
            "synced_at": iso_date(get_value(fund_ws, row, fund_hm, "信息同步日期")),
        })

    updated_dates = [f["synced_at"] for f in funds if f.get("synced_at")]
    fund_document = {
        "schema_version": "1.0.0",
        "updated_at": max(updated_dates) if updated_dates else datetime.now().date().isoformat(),
        "source": "天天基金；由旧版Excel迁移",
        "funds": funds,
    }
    write_text_atomic(funds_path, dump_funds_yaml(fund_document))

    schema = load_schema(target_root / "data" / "schema.json")
    canonical_headers = schema["transactions"]["headers"]
    tx_ws = wb["交易流水"]
    tx_hm = header_map(tx_ws, 5)
    ensure_headers(tx_hm, TX_HEADERS, "交易流水")
    counters: dict[tuple[str, str], int] = defaultdict(int)
    migrated_at = datetime.now().astimezone().isoformat(timespec="seconds")
    batch_id = "MIGRATION-XLSX-20260817"
    transactions: list[dict[str, str]] = []
    confirmed = pending = 0
    for row in range(6, tx_ws.max_row + 1):
        request_date = iso_date(get_value(tx_ws, row, tx_hm, "申请日期"))
        code = code6(get_value(tx_ws, row, tx_hm, "基金代码"))
        tx_type = str(get_value(tx_ws, row, tx_hm, "交易类型") or "").strip()
        if not request_date and not code and not tx_type:
            continue
        confirm_date = iso_date(get_value(tx_ws, row, tx_hm, "确认日期"))
        note = str(get_value(tx_ws, row, tx_hm, "备注") or "").strip()
        status = "待确认" if (not confirm_date or "待确认" in note or "未确认" in note) else "已确认"
        if status == "待确认":
            pending += 1
        else:
            confirmed += 1
        item = {h: "" for h in canonical_headers}
        item.update({
            "交易ID": make_tx_id(request_date, code, counters),
            "申请日期": request_date,
            "确认日期": confirm_date,
            "基金代码": code,
            "交易类型": tx_type,
            "申请成交金额": decimal_text(get_value(tx_ws, row, tx_hm, "申请/成交金额")),
            "确认份额": decimal_text(get_value(tx_ws, row, tx_hm, "确认份额")),
            "确认净值": decimal_text(get_value(tx_ws, row, tx_hm, "确认净值")),
            "手续费": decimal_text(get_value(tx_ws, row, tx_hm, "手续费")),
            "实际到账分红": decimal_text(get_value(tx_ws, row, tx_hm, "实际到账/分红")),
            "确认状态": status,
            "策略标签": str(get_value(tx_ws, row, tx_hm, "策略标签") or "").strip(),
            "备注": note,
            "关联交易ID": "",
            "录入来源": "历史Excel迁移",
            "录入批次": batch_id,
            "录入时间": migrated_at,
            "最近修改来源": "",
            "最近修改批次": "",
            "最近修改时间": "",
        })
        transactions.append(item)
    write_csv_atomic(tx_path, canonical_headers, transactions)
    return {
        "workbook": str(workbook.resolve()),
        "target_root": str(target_root.resolve()),
        "fund_count": len(funds),
        "transaction_count": len(transactions),
        "confirmed_count": confirmed,
        "pending_count": pending,
        "funds_path": str(funds_path),
        "transactions_path": str(tx_path),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="把旧版基金Excel迁移到独立的YAML/CSV项目")
    parser.add_argument("workbook")
    parser.add_argument("--target-root", required=True)
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()
    result = migrate(Path(args.workbook), Path(args.target_root), args.force)
    import json
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
