from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

from common import code6, dump_json, header_map, load_schema


def validate(path: str, schema_path: str | None = None) -> dict:
    schema = load_schema(schema_path)
    wb = load_workbook(path, data_only=False, read_only=False)
    errors: list[str] = []
    warnings: list[str] = []

    for key in ("fund_info", "transactions"):
        cfg = schema[key]
        sheet_name = cfg["sheet"]
        if sheet_name not in wb.sheetnames:
            errors.append(f"缺少工作表：{sheet_name}")
            continue
        ws = wb[sheet_name]
        actual = [ws.cell(cfg["header_row"], col).value for col in range(1, len(cfg["headers"]) + 1)]
        if actual != cfg["headers"]:
            errors.append(f"{sheet_name}表头与Schema不一致：{actual}")

    if schema["fund_info"]["sheet"] in wb.sheetnames:
        ws = wb[schema["fund_info"]["sheet"]]
        hm = header_map(ws, schema["fund_info"]["header_row"])
        for row in range(schema["fund_info"]["data_start_row"], ws.max_row + 1):
            raw = ws.cell(row, hm.get("基金代码", 1)).value
            if raw in (None, ""):
                continue
            if len(code6(raw)) != 6:
                errors.append(f"基金信息第{row}行基金代码不是六位：{raw}")
            for field in schema["fund_info"]["required"]:
                if field in hm and ws.cell(row, hm[field]).value in (None, ""):
                    errors.append(f"基金信息第{row}行缺少必填字段：{field}")
            if any(ws.cell(row, hm.get(field, 1)).value in (None, "") for field in schema["fund_info"]["technical_required"]):
                warnings.append(f"基金信息第{row}行缺少关联ETF/SecID，只能做净值技术面")

            secid_cell = ws.cell(row, hm.get("行情SecID", 1))
            secid_value = secid_cell.value
            if secid_value not in (None, ""):
                secid_text = str(secid_value)
                if not isinstance(secid_value, str):
                    errors.append(f"基金信息第{row}行行情SecID必须存储为文本：{secid_value}")
                if not re.fullmatch(r"[01]\.\d{6}", secid_text):
                    errors.append(f"基金信息第{row}行行情SecID格式错误，应为0/1加六位代码：{secid_text}")
                etf_code = code6(ws.cell(row, hm.get("关联场内ETF代码", 1)).value)
                if re.fullmatch(r"[01]\.\d{6}", secid_text) and etf_code and secid_text.split(".", 1)[1] != etf_code:
                    errors.append(f"基金信息第{row}行行情SecID与关联ETF代码不一致：{secid_text} / {etf_code}")
                if secid_cell.hyperlink and "quote.eastmoney.com" not in str(secid_cell.hyperlink.target or ""):
                    warnings.append(f"基金信息第{row}行行情SecID超链接不是东方财富ETF行情页")

    if schema["transactions"]["sheet"] in wb.sheetnames:
        ws = wb[schema["transactions"]["sheet"]]
        hm = header_map(ws, schema["transactions"]["header_row"])
        d6 = str(ws["D6"].value or "")
        k6 = str(ws["K6"].value or "")
        if "'基金信息'!" not in d6 and "基金信息!" not in d6:
            errors.append("D6基金名称公式必须引用基金信息工作表")
        if not k6.startswith("=IF("):
            errors.append("K6现金流公式缺失")
        expected_last_row = 504
        missing_name_formulas = [row for row in range(cfg["data_start_row"], expected_last_row + 1) if not str(ws.cell(row, hm["基金名称(自动)"]).value or "").startswith("=")]
        missing_cash_formulas = [row for row in range(cfg["data_start_row"], expected_last_row + 1) if not str(ws.cell(row, hm["现金流(自动)"]).value or "").startswith("=")]
        if missing_name_formulas:
            errors.append(f"基金名称自动公式未覆盖至第{expected_last_row}行，首个缺失行：{missing_name_formulas[0]}")
        if missing_cash_formulas:
            errors.append(f"现金流自动公式未覆盖至第{expected_last_row}行，首个缺失行：{missing_cash_formulas[0]}")

    return {
        "status": "OK" if not errors else "ERROR",
        "schema_version": schema.get("version"),
        "workbook": str(Path(path).resolve()),
        "errors": errors,
        "warnings": warnings,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="校验基金分析工作簿与固定Schema是否同步")
    parser.add_argument("workbook")
    parser.add_argument("--schema")
    parser.add_argument("--format", choices=["text", "json"], default="text")
    args = parser.parse_args()
    result = validate(args.workbook, args.schema)
    if args.format == "json":
        print(dump_json(result))
    else:
        print(f"[{result['status']}] Schema {result['schema_version']}")
        for item in result["errors"]:
            print(f"ERROR: {item}")
        for item in result["warnings"]:
            print(f"WARN: {item}")
    return 0 if result["status"] == "OK" else 1


if __name__ == "__main__":
    sys.exit(main())