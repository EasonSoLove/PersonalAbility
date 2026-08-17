from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from common import code6, contains_any, dump_json, header_map, iso_date, load_json, load_schema, number, row_dict


def load_navs(path: str | None) -> dict[str, float]:
    if not path:
        return {}
    raw = load_json(path)
    if isinstance(raw, dict) and "funds" in raw:
        raw = raw["funds"]
    result: dict[str, float] = {}
    if isinstance(raw, list):
        for item in raw:
            if isinstance(item, dict):
                nav = item.get("latest_nav", item.get("nav"))
                if nav is not None:
                    result[code6(item.get("fund_code", item.get("code")))] = number(nav)
    elif isinstance(raw, dict):
        for key, value in raw.items():
            nav = value.get("latest_nav", value.get("nav")) if isinstance(value, dict) else value
            result[code6(key)] = number(nav)
    return result


def analyze(workbook: str, nav_json: str | None = None, schema_path: str | None = None) -> dict[str, Any]:
    schema = load_schema(schema_path)
    cfg = schema["transactions"]
    wb = load_workbook(workbook, data_only=True, read_only=False)
    if cfg["sheet"] not in wb.sheetnames:
        raise ValueError(f"缺少工作表：{cfg['sheet']}")
    ws = wb[cfg["sheet"]]
    headers = header_map(ws, cfg["header_row"])
    missing = [h for h in cfg["headers"] if h not in headers]
    if missing:
        raise ValueError(f"交易流水缺少字段：{missing}")

    transactions: list[dict[str, Any]] = []
    pending: list[dict[str, Any]] = []
    warnings: list[str] = []
    label_stats: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    conversion_in = conversion_out = 0

    for row in range(cfg["data_start_row"], ws.max_row + 1):
        item = row_dict(ws, row, headers)
        if all(item.get(k) in (None, "") for k in ("申请日期", "基金代码", "交易类型")):
            continue
        item["row"] = row
        item["基金代码"] = code6(item.get("基金代码"))
        item["申请日期"] = iso_date(item.get("申请日期"))
        item["确认日期"] = iso_date(item.get("确认日期"))
        tx_type = str(item.get("交易类型") or "").strip()
        item["交易类型"] = tx_type
        label = str(item.get("策略标签") or "").strip()
        if label:
            label_stats[label][tx_type] += 1
        if tx_type == "转换转入":
            conversion_in += 1
        if tx_type == "转换转出":
            conversion_out += 1
        is_pending = contains_any(item.values(), cfg["pending_markers"]) or (tx_type in cfg["confirmed_out"] and not item["确认日期"])
        if is_pending:
            pending.append(item)
        else:
            transactions.append(item)

    transactions.sort(key=lambda x: (x.get("确认日期") or x.get("申请日期") or "9999-99-99", x["row"]))

    buy_dates: dict[str, list[date]] = defaultdict(list)
    for item in transactions:
        if item["交易类型"] in cfg["confirmed_in"] and item.get("确认日期"):
            try:
                buy_dates[item["基金代码"]].append(date.fromisoformat(item["确认日期"]))
            except ValueError:
                pass
    for code, dates in buy_dates.items():
        dates.sort()
        if any((dates[i] - dates[i - 2]).days <= 7 for i in range(2, len(dates))):
            warnings.append(f"{code}存在7日内至少3笔确认买入，需核对是否短期快速建仓或越跌越密集")
    navs = load_navs(nav_json)
    holdings: dict[str, dict[str, float]] = defaultdict(lambda: {"shares": 0.0, "cost": 0.0, "realized_pnl": 0.0, "dividends": 0.0, "external_cash": 0.0})

    for item in transactions:
        code = item["基金代码"]
        tx_type = item["交易类型"]
        amount = number(item.get("申请/成交金额"))
        shares = number(item.get("确认份额"))
        nav = number(item.get("确认净值"))
        fee = number(item.get("手续费"))
        received = number(item.get("实际到账/分红"))
        h = holdings[code]

        if tx_type in cfg["confirmed_in"]:
            if shares <= 0:
                warnings.append(f"第{item['row']}行{code}{tx_type}缺少确认份额，未计入持仓")
                continue
            h["shares"] += shares
            h["cost"] += amount + fee
            if tx_type != "转换转入":
                h["external_cash"] += amount + fee
        elif tx_type in cfg["confirmed_out"]:
            if shares <= 0:
                warnings.append(f"第{item['row']}行{code}{tx_type}缺少确认份额，未处理")
                continue
            if h["shares"] <= 0:
                warnings.append(f"第{item['row']}行{code}{tx_type}前没有可用份额")
                continue
            sold = min(shares, h["shares"])
            avg_cost = h["cost"] / h["shares"] if h["shares"] else 0.0
            removed_cost = sold * avg_cost
            proceeds = received if received > 0 else amount
            if proceeds <= 0 and nav > 0:
                proceeds = sold * nav - fee
                warnings.append(f"第{item['row']}行{code}{tx_type}用份额×净值估算到账")
            elif received <= 0:
                warnings.append(f"第{item['row']}行{code}{tx_type}缺少实际到账，使用成交金额")
            h["shares"] -= sold
            h["cost"] -= removed_cost
            h["realized_pnl"] += proceeds - removed_cost
            if "止盈" in str(item.get("策略标签") or "") and proceeds < removed_cost:
                warnings.append(f"第{item['row']}行{code}标记为止盈，但本次到账低于转出成本，标签与实际收益矛盾")
            if tx_type != "转换转出":
                h["external_cash"] -= proceeds
        elif tx_type in cfg["dividend"]:
            h["dividends"] += received
            h["realized_pnl"] += received
            h["external_cash"] -= received

    if conversion_in != conversion_out:
        warnings.append(f"转换转入{conversion_in}笔、转换转出{conversion_out}笔，记录可能不成对，外部现金流可能失真")
    if "再平衡" in label_stats and sum(label_stats["再平衡"].get(t, 0) for t in cfg["confirmed_out"]) == 0:
        warnings.append("标记为‘再平衡’的记录没有减仓交易，本质可能只是继续加仓")
    if "回撤加仓" in label_stats:
        warnings.append("存在‘回撤加仓’记录；需另行核对是否记录了触发阈值、最大仓位和停止条件")

    rows: list[dict[str, Any]] = []
    for code, h in sorted(holdings.items()):
        if h["shares"] < 1e-8:
            h["shares"] = 0.0
            h["cost"] = max(h["cost"], 0.0)
        nav = navs.get(code)
        market_value = h["shares"] * nav if nav else None
        unrealized = market_value - h["cost"] if market_value is not None else None
        break_even = max(h["cost"] - market_value, 0.0) / market_value if market_value and market_value > 0 else None
        rows.append({
            "fund_code": code,
            "shares": round(h["shares"], 4),
            "remaining_cost": round(h["cost"], 2),
            "average_cost": round(h["cost"] / h["shares"], 6) if h["shares"] else None,
            "latest_nav": nav,
            "market_value": round(market_value, 2) if market_value is not None else None,
            "unrealized_pnl": round(unrealized, 2) if unrealized is not None else None,
            "break_even_rise": round(break_even, 6) if break_even is not None else None,
            "realized_pnl": round(h["realized_pnl"], 2),
            "external_net_cash": round(h["external_cash"], 2),
        })

    active = [r for r in rows if r["shares"] > 0]
    total_cost = sum(r["remaining_cost"] for r in active)
    values_known = active and all(r["market_value"] is not None for r in active)
    total_market = sum(r["market_value"] for r in active) if values_known else None
    total_unrealized = total_market - total_cost if total_market is not None else None
    total_break_even = max(total_cost - total_market, 0) / total_market if total_market and total_market > 0 else None

    return {
        "analysis_mode": "portfolio_plus_technical" if active else "technical_only",
        "workbook": str(Path(workbook).resolve()),
        "transaction_count": len(transactions),
        "pending_count": len(pending),
        "pending_trades": pending,
        "holdings": rows,
        "portfolio": {
            "remaining_cost": round(total_cost, 2),
            "market_value": round(total_market, 2) if total_market is not None else None,
            "unrealized_pnl": round(total_unrealized, 2) if total_unrealized is not None else None,
            "break_even_rise": round(total_break_even, 6) if total_break_even is not None else None,
        },
        "warnings": warnings,
    }


def markdown(result: dict[str, Any]) -> str:
    lines = [f"# 交易流水计算", "", f"- 分析分支：{result['analysis_mode']}", f"- 已确认交易：{result['transaction_count']}", f"- 待确认交易：{result['pending_count']}", ""]
    if result["holdings"]:
        lines += ["|基金代码|份额|剩余成本|均价|最新净值|市值|浮动盈亏|回本所需涨幅|", "|---|---:|---:|---:|---:|---:|---:|---:|"]
        for r in result["holdings"]:
            pct = "" if r["break_even_rise"] is None else f"{r['break_even_rise']:.2%}"
            lines.append(f"|{r['fund_code']}|{r['shares']:.4f}|{r['remaining_cost']:.2f}|{r['average_cost'] or ''}|{r['latest_nav'] or ''}|{r['market_value'] or ''}|{r['unrealized_pnl'] or ''}|{pct}|")
    if result["warnings"]:
        lines += ["", "## 警告"] + [f"- {x}" for x in result["warnings"]]
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description="按固定规则计算基金交易流水、持仓成本和盈亏")
    parser.add_argument("workbook")
    parser.add_argument("--nav-json")
    parser.add_argument("--schema")
    parser.add_argument("--format", choices=["markdown", "json"], default="markdown")
    parser.add_argument("--output")
    args = parser.parse_args()
    result = analyze(args.workbook, args.nav_json, args.schema)
    text = dump_json(result) if args.format == "json" else markdown(result)
    if args.output:
        Path(args.output).write_text(text + "\n", encoding="utf-8")
    print(text)
    return 0


if __name__ == "__main__":
    sys.exit(main())