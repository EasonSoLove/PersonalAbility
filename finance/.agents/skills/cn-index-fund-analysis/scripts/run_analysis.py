from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from common import code6, dump_json, header_map, load_schema
from portfolio_analysis import analyze as analyze_portfolio, markdown as portfolio_markdown
from technical_analysis import analyze as analyze_technical, markdown as technical_markdown
from validate_workbook import validate


def workbook_fund_codes(workbook: str, schema_path: str | None = None) -> list[str]:
    schema = load_schema(schema_path)
    wb = load_workbook(workbook, data_only=True, read_only=False)
    codes: list[str] = []
    seen: set[str] = set()
    for key, field in (("fund_info", "基金代码"), ("transactions", "基金代码")):
        cfg = schema[key]
        if cfg["sheet"] not in wb.sheetnames:
            continue
        ws = wb[cfg["sheet"]]
        hm = header_map(ws, cfg["header_row"])
        col = hm.get(field)
        if not col:
            continue
        for row in range(cfg["data_start_row"], ws.max_row + 1):
            code = code6(ws.cell(row, col).value)
            if len(code) == 6 and code.isdigit() and code not in seen:
                seen.add(code)
                codes.append(code)
    return codes


def run(workbook: str, mapping: str, nav_json: str | None = None, bars_json: str | None = None, nav_bars_json: str | None = None, end: str | None = None, schema_path: str | None = None) -> dict[str, Any]:
    validation = validate(workbook, schema_path)
    if validation["status"] != "OK":
        return {
            "status": "ERROR",
            "workbook": str(Path(workbook).resolve()),
            "validation": validation,
            "analysis_mode": None,
            "portfolio": None,
            "technical": None,
        }
    portfolio = analyze_portfolio(workbook, nav_json, schema_path)
    codes = workbook_fund_codes(workbook, schema_path)
    technical = analyze_technical(mapping, codes, end=end, bars_json=bars_json, nav_bars_json=nav_bars_json)
    return {
        "status": "OK",
        "workbook": str(Path(workbook).resolve()),
        "analysis_mode": portfolio["analysis_mode"],
        "validation": validation,
        "portfolio": portfolio,
        "technical": technical,
    }


def markdown(result: dict[str, Any]) -> str:
    if result["status"] != "OK":
        lines = ["# 基金组合分析", "", "模板校验未通过，未继续计算。"]
        lines += [f"- {x}" for x in result["validation"]["errors"]]
        return "\n".join(lines)
    branch = "交易流水 + 技术面" if result["analysis_mode"] == "portfolio_plus_technical" else "仅技术面"
    parts = ["# 基金组合分析", "", f"- 分析分支：{branch}", f"- 工作簿：{result['workbook']}", "", portfolio_markdown(result["portfolio"]), "", technical_markdown(result["technical"])]
    return "\n".join(parts)


def main() -> int:
    parser = argparse.ArgumentParser(description="校验模板并按是否持仓自动选择分析分支")
    parser.add_argument("workbook")
    parser.add_argument("--mapping", required=True)
    parser.add_argument("--nav-json")
    parser.add_argument("--bars-json")
    parser.add_argument("--nav-bars-json")
    parser.add_argument("--end")
    parser.add_argument("--schema")
    parser.add_argument("--format", choices=["markdown", "json"], default="markdown")
    parser.add_argument("--output")
    args = parser.parse_args()
    result = run(args.workbook, args.mapping, args.nav_json, args.bars_json, args.nav_bars_json, args.end, args.schema)
    text = dump_json(result) if args.format == "json" else markdown(result)
    if args.output:
        Path(args.output).write_text(text + "\n", encoding="utf-8")
    print(text)
    return 0 if result["status"] == "OK" else 1


if __name__ == "__main__":
    sys.exit(main())
